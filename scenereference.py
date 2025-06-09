import os
import requests
import pandas as pd
import subprocess
import re
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage



# Define paths
excel_file = 'scan_results.xlsx'
#input_file_name = "resource/hd2023.csv"
#output_file_name = "resource/crawler.csv"
results_dir = 'results'

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Scene Screenshots"

# Set headers
ws.append(["Folder Name", "Screenshot", "Scene Number"])

# Row counter for inserting images
row = 2

# Traverse the results directory
for folder_name in os.listdir(results_dir):
    folder_path = os.path.join(results_dir, folder_name)
    if os.path.isdir(folder_path):
        print(f"Processing {folder_path}...")
        scenes_path = os.path.join(folder_path, 'scenes')
        if os.path.exists(scenes_path):
            for file in os.listdir(scenes_path):
                if file.startswith("scene_") and file.endswith("_screenshot.jpg"):
                    print(f"Processing {file}...")
                    match = re.search(r"scene_(\d+)_screenshot\.jpg", file)
                    scene_number = int(match.group(1))
                    print(f"Scene number should be: {scene_number}...")
                    image_path = os.path.join(scenes_path, file)
                    #img = Image(image_path)
                    # Resize image to thumbnail
                    if os.path.exists(os.path.join(scenes_path, f"thumb_{file}")):
                        thumb_path = os.path.join(scenes_path, f"thumb_{file}")
                    else:
                        with PILImage.open(image_path) as img:
                            img.thumbnail((125, 125))
                            thumb_path = os.path.join(scenes_path, f"thumb_{file}")
                            img.save(thumb_path)
                    ws.cell(row=row, column=1, value=folder_name)
                    img = Image(thumb_path)
                    ws.add_image(img, f'B{row}')
                    ws.cell(row=row, column=3, value=scene_number)
                    row += 1  # Leave space for image height

# Save the workbook
wb.save("scene_screenshots.xlsx")
print("Spreadsheet 'scene_screenshots.xlsx' created successfully.")
